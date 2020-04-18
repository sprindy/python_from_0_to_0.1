#!/usr/local/bin/python3

import re

import pandas as pd
from matplotlib import pyplot as plt
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

# 指定表格列标题
header = ['head', 'null', 'time', 'null', 'a_sample', 'index', 'TS', 'ts', 'LSB',
          'AX', 'AY', 'AZ', 'status', 'null', 'line_num', 'null', 'file_name']
# 用上面的标题创建二位数组（表格）
sensor_data = pd.DataFrame(columns=header)
# print(sensor_data)

# 给变量赋值
log_file = 'log.txt'
log_exelc_file = 'ts.xlsx'

# 利用python自带的工具绘图
def plot_outside_excel():
    key_words = "a_sample"
    a_sample_log = []
    # 打印列表的长度（这时候是0，因为是空的）
    print(len(a_sample_log))

    # 只读方式打开日志文件
    with open(log_file, 'r') as f:
        # 读入所有行放到lines变量里
        lines = f.readlines()
        # print(lines)
        # 对每一行进行处理
        for line in lines:
            # print(line)
            # 在行里搜索关键字
            m = re.search(key_words, line)
            # 搜到带关键字的行
            if m is not None:
                # print(line)
                # 对搜到的行进行分割，分割符是空格/冒号/逗号，字符串的一次只能用一个
                # a_sample_log = line.split(" ")
                # 对搜到的行进行分割，分割符是空格/冒号/逗号，三个可以一起用
                a_sample_log = re.split(' |, |: ', line)
                # print(len(a_sample_log))
                # print(a_sample_log)
                # 把分割后的这个列表逐行放到表格里的行里， len是求表格的行树，每新增一行，它的值都会增加
                sensor_data.loc[len(sensor_data)] = a_sample_log


    # 打印表格的头几行
    # print(sensor_data.head(3))
    # 打印表格的更多信息
    print(sensor_data.describe())
    # 保存数组到表格文件里去
    sensor_data.to_excel(log_exelc_file)

    # 读出表格文件里的数据到数组里
    reliable_data = pd.read_excel(log_exelc_file)
    # 对第10列求标准差
    print(reliable_data.iloc[:, 10].std())

    # 绘图，条形图，点状图，横坐标是第几个数，纵坐标是数值
    # plt.bar(reliable_data['index'], reliable_data['AX'])
    plt.scatter(reliable_data['index'], reliable_data['AX'])
    # 加标题和坐标标签
    plt.title('realiable data samples:')
    plt.xlabel('AX')
    plt.ylabel('Index')
    plt.show()

# 在表格里面插入散点图
def plot_inside_excel():
    # wb = Workbook()
    wb = load_workbook(filename = log_exelc_file)
    ws = wb.active

    chart = ScatterChart()
    chart.title = "Scatter Chart"
    chart.style = 12
    chart.x_axis.title = 'Index'
    chart.y_axis.title = 'Real sample value of X'
    # 增加行列的高度和宽度
    chart.height = chart.height + 8
    chart.width = chart.width + 24

    # 横坐标
    xvalues = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    # 纵坐标
    values = Reference(ws, min_col=11, min_row=2, max_row=ws.max_row)
    # 序列1
    series = Series(values, xvalues, title="serial 1", title_from_data=False)
    chart.series.append(series)

    # 从A10插入图表
    ws.add_chart(chart, "A10")

    wb.save(log_exelc_file)

# plot_outside_excel()
plot_inside_excel()
